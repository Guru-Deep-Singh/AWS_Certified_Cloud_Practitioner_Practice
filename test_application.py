import re
import os
from pathlib import Path

import pandas as pd
import streamlit as st
import json
from streamlit.components.v1 import html as st_html
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------
# Copy button helper (added)
# -------------------------

def copy_to_clipboard_button(text: str, key: str, label: str = "📋 Copy"):
    """
    Render a tiny HTML button that copies `text` to clipboard.
    - Uses dark theme styling (black background to match Streamlit dark).
    - Fully rounded pill border so the bottom edge doesn't look clipped.
    - On success, label changes to "Copied" briefly.
    """
    js_safe_text = json.dumps(text)
    button_id = f"{key}_btn"
    status_id = f"{key}_status"
    html_snippet = """
        <div style="display:flex;align-items:center;gap:8px;justify-content:flex-end;">
          <button id="{button_id}" style="
              display:inline-block;
              font-size:0.85rem;
              padding:0.38rem 0.9rem;
              line-height:1.1;
              border-radius:9999px;
              border:1px solid #3a3a3a;
              background:#0e1117;
              color:#ffffff;
              cursor:pointer;
              outline:none;
          ">
            {label}
          </button>
          <span id="{status_id}" style="font-size:0.8rem;opacity:0.7;"></span>
        </div>
        <script>
        (function(){{
          const btn = document.getElementById("{button_id}");
          const status = document.getElementById("{status_id}");
          const text = {js_safe_text};
          const original = btn.textContent;

          function fallbackCopy() {{
            const ta = document.createElement('textarea');
            ta.value = text;
            ta.setAttribute('readonly', '');
            ta.style.position = 'fixed';
            ta.style.left = '-9999px';
            document.body.appendChild(ta);
            ta.select();
            let ok = false;
            try {{ ok = document.execCommand('copy'); }} catch (e) {{ ok = false; }}
            document.body.removeChild(ta);
            return ok;
          }}

          async function doCopy() {{
            let ok = false;
            btn.disabled = true;
            try {{
              if (navigator.clipboard && window.isSecureContext) {{
                await navigator.clipboard.writeText(text);
                ok = true;
              }}
            }} catch (e) {{
              ok = false;
            }}
            if (!ok) ok = fallbackCopy();

            if (ok) {{
              btn.textContent = "Copied";
              status.textContent = "";
            }} else {{
              status.textContent = "Copy failed";
              status.style.opacity = 1;
            }}
            setTimeout(() => {{
              btn.textContent = original;
              btn.disabled = false;
              status.textContent = "";
            }}, 1500);
          }}
          btn.addEventListener('click', doCopy);
        }})();
        </script>
    """.format(button_id=button_id, status_id=status_id, label=label, js_safe_text=js_safe_text)
    # Increase component height a bit to avoid clipping
    st_html(html_snippet, height=56)

# ===================== Page setup (must be the first Streamlit call) =====================
APP_DIR = Path(__file__).parent.resolve()
LOGO_PATH = APP_DIR / "images" / "aws_ccp.png"  # your image path

st.set_page_config(
    page_title="AWS Certified Cloud Practitioner: Practice Tests",
    page_icon=str(LOGO_PATH),  # favicon in browser tab
)

# ===================== File Parsing Utilities =====================

def parse_md_file(md_path: str):
    with open(md_path, 'r', encoding='utf-8') as file:
        content = "\n".join(line for line in file if not line.strip().startswith("#"))

    # Split on lines that start with "<number>. "
    question_blocks = re.split(r'\n(?=\d+\.\s)', content)

    questions = []
    for block in question_blocks:
        if not block.strip():
            continue

        lines = block.strip().split('\n')
        q_text = lines[0]
        options = []
        for line in lines[1:]:
            line = line.strip()
            # lines look like "- A. Option text"
            if re.match(r'-\s+[A-Z]\.', line):
                options.append(line[2:].strip())  # keep "A. ..." prefix

        # heuristic: multi-select if prompt mentions "Choose/Select TWO/THREE/FOUR"
        multi_select = bool(re.search(r'\b(Choose|Select)\b.*\b(TWO|THREE|FOUR)\b', q_text, re.IGNORECASE))

        questions.append({
            'question': q_text,
            'options': options,
            'multi_select': multi_select
        })

    return questions


def parse_answers_file(answers_path: str):
    with open(answers_path, 'r', encoding='utf-8') as file:
        lines = [line for line in file if not line.strip().startswith("#")]

    answers = {}
    for line in lines:
        if line.strip() and re.match(r'\d+\.\s', line):
            q_num, ans = line.strip().split('.', 1)
            ans_list = [a.strip() for a in ans.strip().split(',')]
            answers[q_num] = ans_list

    return answers


def extract_option_letter(option_text: str):
    """Extract 'A' from 'A. Something'."""
    match = re.match(r'^([A-Z])\.', option_text.strip())
    if match:
        return match.group(1)
    return option_text.strip()


def extract_test_number(filename: str):
    match = re.search(r'(\d+)(?=\.md$)', filename)
    return int(match.group(1)) if match else float('inf')


def grade_one(user_sel, correct_list):
    """Return (yours_text, correct_text, is_correct) for a question."""
    correct_text = ", ".join(correct_list)
    if isinstance(user_sel, list):
        selected_letters = [extract_option_letter(a) for a in (user_sel or [])]
        yours_text = ", ".join(selected_letters) if selected_letters else ""
        is_correct = set(a.upper() for a in selected_letters) == set(c.upper() for c in correct_list)
    else:
        sel = extract_option_letter(user_sel) if user_sel else ""
        yours_text = sel
        is_correct = ({sel.upper()} if sel else set()) == set(c.upper() for c in correct_list)
    return yours_text, correct_text, is_correct


# ===================== Streamlit App =====================

def main():
    # ---------- Header with AWS badge image and aligned titles ----------
    col_logo, col_title = st.columns([1, 6])
    with col_logo:
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width=90)
        else:
            st.warning(f"Logo not found at: {LOGO_PATH}")
    with col_title:
        # Pull the title block slightly up to vertically center with the badge
        st.markdown(
            """
            <div style="margin-top:-15px;">
              <h2 style="margin:0; line-height:1.15;">AWS Certified Cloud Practitioner: Practice Tests</h2>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # Add extra vertical space before the select label
    st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)

    # ---------- Paths ----------
    question_dir = "output_md_files"
    answer_dir = "answers_md_files"

    available_tests = sorted([
        f for f in os.listdir(question_dir)
        if f.endswith(".md") and os.path.exists(os.path.join(answer_dir, f))
    ], key=extract_test_number)

    test_selected = st.selectbox("Select a test to take:", available_tests)

    if test_selected:
        md_path = os.path.join(question_dir, test_selected)
        ans_path = os.path.join(answer_dir, test_selected)

        try:
            questions = parse_md_file(md_path)
            correct_answers = parse_answers_file(ans_path)
        except Exception as e:
            st.error(f"❌ Error parsing selected files: {e}")
            return

        st.success(f"Parsed {len(questions)} questions from **{test_selected}**!")

        submitted = st.session_state.get("submitted", False)
        user_answers = {}
        flags = {}

        # ---------- Render questions: left inputs, right correctness panel ----------
        for idx, q in enumerate(questions, 1):
            idx_str = str(idx)

            st.markdown("---")
            st.markdown(f"**{q['question']}**")

            col_left, col_right = st.columns([3, 2])

            with col_left:
                if q['multi_select']:
                    selected = st.multiselect(
                        "Select your answers:",
                        q['options'],
                        key=f"answer_{idx}",
                    )
                else:
                    selected = st.radio(
                        "Select your answer:",
                        q['options'],
                        key=f"answer_{idx}",
                        index=None  # don't preselect
                    )

                chk_col, copy_col = st.columns([5, 2])

                with chk_col:

                    flag = st.checkbox("🚩 Flag this question", key=f"flag_{idx}")

                with copy_col:

                    copy_text = q['question'] + "\n\n" + "\n".join([f"- {opt}" for opt in q['options']])

                    copy_to_clipboard_button(copy_text, key=f"copy_{idx}", label="Copy Question!")

                user_answers[idx_str] = selected
                flags[idx_str] = flag

            with col_right:
                if submitted:
                    yours_text, correct_text, is_correct = grade_one(selected, correct_answers.get(idx_str, []))
                    badge = ":green-background[✅ Correct]" if is_correct else ":red-background[❌ Incorrect]"
                    st.markdown(badge)
                    st.markdown(
                        f"- **Your Answer(s):** {yours_text or '_No selection_'}  \n"
                        f"- **Correct Answer(s):** {correct_text or '_N/A_'}"
                    )
                else:
                    st.caption("Submit to reveal correct answers →")

        # ---------- Submit: grade + export (inputs remain editable) ----------
        if st.button("Submit Test ✅"):
            records = []
            correct_count = 0
            for idx in user_answers.keys():
                yours_text, correct_text, is_correct = grade_one(user_answers[idx], correct_answers.get(idx, []))
                result = "Correct" if is_correct else "Incorrect"
                if is_correct:
                    correct_count += 1
                records.append({
                    "Question Number": idx,
                    "Your Answer(s)": yours_text,
                    "Correct Answer(s)": correct_text,
                    "Result": result,
                    "Flagged": "FLAGGED" if flags.get(idx) else ""
                })

            st.session_state["submitted"] = True

            # Excel export with red/green row highlighting
            df = pd.DataFrame(records)
            output_filename = test_selected.replace(".md", ".xlsx")
            df.to_excel(output_filename, index=False)

            wb = load_workbook(output_filename)
            ws = wb.active
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if row[3].value == "Incorrect":
                    for cell in row:
                        cell.fill = red_fill
                elif row[3].value == "Correct":
                    for cell in row:
                        cell.fill = green_fill

            wb.save(output_filename)

            total_questions = len(questions)
            score_percentage = (correct_count / total_questions) * 100 if total_questions else 0
            st.session_state["score_msg"] = f"You scored {correct_count} out of {total_questions} ({score_percentage:.2f}%) 🎯"
            st.session_state["excel_path"] = output_filename

            # Immediate rerun so answers reveal right away
            try:
                st.rerun()
            except Exception:
                pass

        # ---------- Submitted view: score + single keyed download button ----------
        if st.session_state.get("submitted"):
            st.success(st.session_state.get("score_msg", "Graded."))
            excel_path = st.session_state.get("excel_path")
            if excel_path and os.path.exists(excel_path):
                with open(excel_path, "rb") as f:
                    st.download_button(
                        "Download your graded Excel file 📄",
                        data=f,
                        file_name=os.path.basename(excel_path),
                        key=f"download_{os.path.basename(excel_path)}"
                    )

        # ---------- Hide correct answers/score (keep selections) ----------
        if st.session_state.get("submitted"):
            if st.button("🔄 Hide correct answers / score"):
                st.session_state["submitted"] = False
                st.session_state.pop("score_msg", None)
                st.session_state.pop("excel_path", None)
                try:
                    st.rerun()
                except Exception:
                    pass


# ===================== Run App =====================
if __name__ == "__main__":
    main()

